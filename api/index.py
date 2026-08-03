import base64
import json
from datetime import datetime
from io import BytesIO

import pandas as pd
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


app = Flask(__name__, template_folder="../templates", static_folder="../static")
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024
app.json.sort_keys = False

DOMAIN_ORDER = [
    "Cybersecurity",
    "Quantum Computing",
    "Cloud Computing",
    "Artificial Intelligence",
]

GENERATED_COLUMNS = [
    "Applicant Number",
    "Grouping",
    "Event Sequence",
    "Subdomain Allocation 1",
    "Subdomain Allocation 2",
    "Allocation Status",
    "Attending",
]

# Master batch rule: matched existing applicants keep these generated fields exactly.
PRESERVED_ALLOCATION_COLUMNS = GENERATED_COLUMNS

OPTION_COLUMN_ALIASES = {
    "first": ["1stoption", "1stchoice", "firstoption", "firstchoice", "option1", "choice1"],
    "second": ["2ndoption", "2ndchoice", "secondoption", "secondchoice", "option2", "choice2"],
    "third": ["3rdoption", "3rdchoice", "thirdoption", "thirdchoice", "option3", "choice3"],
}

PREFERENCE_ALIASES = {
    "cybersecurity": "Cybersecurity",
    "cybersec": "Cybersecurity",
    "informationsecurity": "Cybersecurity",
    "infosec": "Cybersecurity",
    "quantumcomputing": "Quantum Computing",
    "quantum": "Quantum Computing",
    "cloudcomputing": "Cloud Computing",
    "cloud": "Cloud Computing",
    "artificialintelligence": "Artificial Intelligence",
    "ai": "Artificial Intelligence",
}


def compact_key(value):
    return "".join(ch for ch in str(value).lower() if ch.isalnum())


def normalize_preference(value):
    return PREFERENCE_ALIASES.get(compact_key(value))


def detect_option_columns(columns):
    compact_columns = {compact_key(column): column for column in columns}
    suggestions = {"first": "", "second": "", "third": ""}
    for field, aliases in OPTION_COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in compact_columns:
                suggestions[field] = compact_columns[alias]
                break
    return suggestions


def load_sheet(file_storage, sheet_name=None):
    if not file_storage or not file_storage.filename.lower().endswith(".xlsx"):
        raise ValueError("Please upload a valid .xlsx workbook.")

    data = BytesIO(file_storage.read())
    try:
        excel_file = pd.ExcelFile(data, engine="openpyxl")
    except Exception as exc:
        raise ValueError("The uploaded workbook could not be opened.") from exc

    if not excel_file.sheet_names:
        raise ValueError("The workbook does not contain any worksheets.")

    selected_sheet = sheet_name or excel_file.sheet_names[0]
    if selected_sheet not in excel_file.sheet_names:
        raise ValueError("The selected worksheet was not found in the workbook.")

    df = pd.read_excel(excel_file, sheet_name=selected_sheet, dtype=str, keep_default_na=False, engine="openpyxl")
    df.columns = [str(column).strip() for column in df.columns]
    non_blank = df.apply(lambda row: any(str(value).strip() for value in row), axis=1)
    return excel_file.sheet_names, selected_sheet, df.loc[non_blank].copy()


def load_previous_allocations(file_storage, sheet_name=None):
    if not file_storage or not file_storage.filename:
        return None
    if not file_storage.filename.lower().endswith(".xlsx"):
        raise ValueError("Please upload a valid previous .xlsx allocation workbook.")

    data = BytesIO(file_storage.read())
    try:
        excel_file = pd.ExcelFile(data, engine="openpyxl")
    except Exception as exc:
        raise ValueError("The previous allocation workbook could not be opened.") from exc

    selected_sheet = sheet_name or ("Allocated Applicants" if "Allocated Applicants" in excel_file.sheet_names else excel_file.sheet_names[0])
    if selected_sheet not in excel_file.sheet_names:
        raise ValueError("The selected previous worksheet was not found.")
    df = pd.read_excel(excel_file, sheet_name=selected_sheet, dtype=str, keep_default_na=False, engine="openpyxl")
    df.columns = [str(column).strip() for column in df.columns]
    missing_columns = [column for column in GENERATED_COLUMNS if column not in df.columns]
    if missing_columns:
        raise ValueError("The previous allocation workbook must contain the generated allocation columns.")
    non_blank = df.apply(lambda row: any(str(value).strip() for value in row), axis=1)
    return df.loc[non_blank].copy()


def inspect_previous_allocation(file_storage, sheet_name=None):
    if not file_storage or not file_storage.filename.lower().endswith(".xlsx"):
        raise ValueError("Please upload a valid previous .xlsx allocation workbook.")
    data = BytesIO(file_storage.read())
    try:
        excel_file = pd.ExcelFile(data, engine="openpyxl")
    except Exception as exc:
        raise ValueError("The previous allocation workbook could not be opened.") from exc
    selected_sheet = sheet_name or ("Allocated Applicants" if "Allocated Applicants" in excel_file.sheet_names else excel_file.sheet_names[0])
    if selected_sheet not in excel_file.sheet_names:
        raise ValueError("The selected previous worksheet was not found.")
    df = pd.read_excel(excel_file, sheet_name=selected_sheet, dtype=str, keep_default_na=False, engine="openpyxl")
    df.columns = [str(column).strip() for column in df.columns]
    non_blank = df.apply(lambda row: any(str(value).strip() for value in row), axis=1)
    missing_columns = [column for column in GENERATED_COLUMNS if column not in df.columns]
    return {
        "worksheets": excel_file.sheet_names,
        "selectedWorksheet": selected_sheet,
        "applicantCount": int(non_blank.sum()),
        "hasGeneratedColumns": not missing_columns,
        "missingColumns": missing_columns,
    }


def validate_config(max_applicants, total_mentors, mentor_counts):
    if max_applicants <= 0:
        raise ValueError("Maximum applicants must be a positive integer.")
    if total_mentors <= 0:
        raise ValueError("Total mentors must be a positive integer.")
    if any(value < 0 for value in mentor_counts.values()):
        raise ValueError("Mentor counts cannot be negative.")
    if sum(mentor_counts.values()) != total_mentors:
        raise ValueError("The four subdomain mentor counts must equal total mentors.")
    if sum(1 for value in mentor_counts.values() if value > 0) < 2:
        raise ValueError("At least two subdomains must have one or more mentors.")


def applicant_name(row):
    for columns in [("Given Name", "Surname"), ("First Name", "Last Name")]:
        parts = [str(row.get(column, "")).strip() for column in columns if str(row.get(column, "")).strip()]
        if parts:
            return " ".join(parts)
    for column in ["Name", "Full Name"]:
        if str(row.get(column, "")).strip():
            return str(row.get(column, "")).strip()
    return ""


def find_column(columns, aliases):
    compact_columns = {compact_key(column): column for column in columns}
    for alias in aliases:
        if alias in compact_columns:
            return compact_columns[alias]
    return None


def identity_key(row):
    mobile_column = find_column(row.index, ["mobilenumber", "mobile", "phone", "contactnumber", "contact"])
    email_column = find_column(row.index, ["personalemail", "email", "emailaddress"])
    given_column = find_column(row.index, ["givenname", "firstname"])
    surname_column = find_column(row.index, ["surname", "lastname", "familyname"])
    mobile = compact_key(row.get(mobile_column, "")) if mobile_column else ""
    email = str(row.get(email_column, "")).strip().lower() if email_column else ""
    given_name = compact_key(row.get(given_column, "")) if given_column else ""
    surname = compact_key(row.get(surname_column, "")) if surname_column else ""
    if not all([surname, given_name, mobile, email]):
        return None
    return (surname, given_name, mobile, email)


def previous_allocation_lookup(previous_df):
    if previous_df is None:
        return {}, []

    lookup = {}
    warnings = []
    for index, row in previous_df.iterrows():
        record = row.to_dict()
        key = identity_key(row)
        if not key:
            warnings.append(
                {
                    "Applicant Number": str(row.get("Applicant Number", "")),
                    "Applicant Name": applicant_name(row),
                    "Excel Row Number": int(index) + 2,
                    "Warning Type": "Missing identifier",
                    "Warning Description": "Previous allocation row is missing surname, given name, mobile number, or personal email for matching.",
                }
            )
            continue
        if key in lookup:
            raise ValueError("The previous allocation workbook contains duplicate applicant identifiers.")
        lookup[key] = record
    return lookup, warnings


def match_previous_record(row, lookup):
    key = identity_key(row)
    if not key:
        return None
    return lookup.get(key)


def make_warning(applicant, warning_type, description):
    return {
        "Applicant Number": applicant["applicant_number"],
        "Applicant Name": applicant["name"],
        "Excel Row Number": applicant["excel_row_number"],
        "Warning Type": warning_type,
        "Warning Description": description,
    }


def build_applicants(df, mappings):
    for key in ["first", "second", "third"]:
        if not mappings.get(key) or mappings[key] not in df.columns:
            raise ValueError("Please map all three option columns.")
    if len(set(mappings.values())) != 3:
        raise ValueError("Each option field must use a different Excel column.")

    applicants = []
    warnings = []
    for applicant_number, (index, row) in enumerate(df.iterrows(), start=1):
        applicant = {
            "applicant_number": applicant_number,
            "excel_index": index,
            "excel_row_number": int(index) + 2,
            "name": applicant_name(row),
            "group": "Group 1" if applicant_number % 2 == 1 else "Group 2",
            "preferences": [],
        }
        seen = set()
        for key in ["first", "second", "third"]:
            raw_value = str(row.get(mappings[key], "")).strip()
            normalized = normalize_preference(raw_value)
            if not raw_value:
                warnings.append(make_warning(applicant, "Blank preference", f"{mappings[key]} is blank."))
            elif not normalized:
                warnings.append(make_warning(applicant, "Invalid preference", f"{raw_value} is not a recognised subdomain."))
            elif normalized in seen:
                warnings.append(make_warning(applicant, "Duplicate preference", f"{normalized} was selected more than once."))
            else:
                seen.add(normalized)
                applicant["preferences"].append(normalized)
        applicants.append(applicant)
    return applicants, warnings


def choose_fallback(allocated, usage, capacity):
    candidates = []
    for domain in DOMAIN_ORDER:
        if domain in allocated or capacity[domain] <= 0 or usage[domain] >= capacity[domain]:
            continue
        remaining_percent = (capacity[domain] - usage[domain]) / capacity[domain]
        candidates.append((-remaining_percent, DOMAIN_ORDER.index(domain), domain))
    if not candidates:
        return None
    return sorted(candidates)[0][2]


def try_allocate(applicants, mentor_counts, capacity_per_mentor, initial_usage=None):
    capacity = {domain: mentor_counts[domain] * capacity_per_mentor for domain in DOMAIN_ORDER}
    usage = (initial_usage or {domain: 0 for domain in DOMAIN_ORDER}).copy()
    allocations = []

    for applicant in applicants:
        allocated = []
        fallback_slots = 0
        for _slot in range(2):
            chosen = None
            for preference in applicant["preferences"]:
                if preference not in allocated and capacity[preference] > 0 and usage[preference] < capacity[preference]:
                    chosen = preference
                    break
            if chosen is None:
                chosen = choose_fallback(allocated, usage, capacity)
                if chosen:
                    fallback_slots += 1
            if not chosen:
                return None
            usage[chosen] += 1
            allocated.append(chosen)
        allocations.append(
            {
                "applicant_number": applicant["applicant_number"],
                "allocation_1": allocated[0],
                "allocation_2": allocated[1],
                "fallback_slots": fallback_slots,
            }
        )
    return allocations, usage, capacity


def allocate_group(applicants, mentor_counts, starting_capacity_per_mentor=5, initial_usage=None):
    initial_usage = initial_usage or {domain: 0 for domain in DOMAIN_ORDER}
    if not applicants:
        capacity = {domain: mentor_counts[domain] * starting_capacity_per_mentor for domain in DOMAIN_ORDER}
        return allocation_result([], starting_capacity_per_mentor, initial_usage.copy(), capacity, 0, 0, [])

    existing_allocations = sum(initial_usage.values())
    max_needed = max(starting_capacity_per_mentor, existing_allocations + len(applicants) * 2)
    for capacity_per_mentor in range(starting_capacity_per_mentor, max_needed + 1):
        result = try_allocate(applicants, mentor_counts, capacity_per_mentor, initial_usage)
        if not result:
            continue
        allocations, usage, capacity = result
        fallback_count = sum(item["fallback_slots"] for item in allocations)
        warnings = []
        if capacity_per_mentor > starting_capacity_per_mentor:
            for applicant in applicants:
                warnings.append(make_warning(applicant, "Capacity expansion required", f"{applicant['group']} used {capacity_per_mentor} participants per mentor."))
        for item in allocations:
            if item["fallback_slots"]:
                applicant = next(app for app in applicants if app["applicant_number"] == item["applicant_number"])
                warnings.append(make_warning(applicant, "Fallback used", "One or more preferred subdomains were unavailable."))
        return allocation_result(allocations, capacity_per_mentor, usage, capacity, fallback_count, 0, warnings)

    warnings = [make_warning(applicant, "Unable to allocate", "The applicant could not receive two distinct subdomains.") for applicant in applicants]
    empty_usage = {domain: 0 for domain in DOMAIN_ORDER}
    final_capacity = {domain: mentor_counts[domain] * max_needed for domain in DOMAIN_ORDER}
    return allocation_result([], max_needed, empty_usage, final_capacity, 0, len(applicants), warnings)


def allocation_result(allocations, final_limit, usage, capacity, fallback_count, unallocated_count, warnings):
    return {
        "allocations": allocations,
        "final_capacity_per_mentor": final_limit,
        "usage": usage,
        "capacity": capacity,
        "remaining": {domain: capacity[domain] - usage[domain] for domain in DOMAIN_ORDER},
        "fallback_count": fallback_count,
        "unallocated_count": unallocated_count,
        "warnings": warnings,
    }


def generated_values_from_previous(record):
    return {column: str(record.get(column, "")).strip() for column in PRESERVED_ALLOCATION_COLUMNS}


def usage_from_existing(existing_records):
    usage = {
        "Group 1": {domain: 0 for domain in DOMAIN_ORDER},
        "Group 2": {domain: 0 for domain in DOMAIN_ORDER},
    }
    for record in existing_records:
        group = str(record.get("Grouping", "")).strip()
        if group not in usage:
            continue
        for column in ["Subdomain Allocation 1", "Subdomain Allocation 2"]:
            domain = str(record.get(column, "")).strip()
            if domain in DOMAIN_ORDER:
                usage[group][domain] += 1
    return usage


def build_existing_capacity_warnings(existing_usage, mentor_counts, capacity_per_mentor=5):
    warnings = []
    for group, usage in existing_usage.items():
        for domain in DOMAIN_ORDER:
            expected_capacity = mentor_counts[domain] * capacity_per_mentor
            over_by = usage[domain] - expected_capacity
            if over_by > 0:
                warnings.append(
                    {
                        "group": group,
                        "domain": domain,
                        "usage": usage[domain],
                        "capacity": expected_capacity,
                        "over_by": over_by,
                    }
                )
    return warnings


def process_allocation(df, mappings, max_applicants, total_mentors, mentor_counts, previous_df=None):
    validate_config(max_applicants, total_mentors, mentor_counts)
    if df.empty:
        raise ValueError("The selected worksheet does not contain any applicants.")
    if len(df) > max_applicants:
        raise ValueError(f"The workbook contains {len(df)} applicants, above the configured maximum of {max_applicants}.")

    applicants, warnings = build_applicants(df, mappings)
    previous_lookup, previous_warnings = previous_allocation_lookup(previous_df)
    warnings.extend(previous_warnings)

    existing_records = []
    new_applicants = []
    allocation_lookup = {}
    next_new_number = len(previous_df) if previous_df is not None else 0

    for applicant in applicants:
        row = df.loc[applicant["excel_index"]]
        previous_record = match_previous_record(row, previous_lookup)
        if previous_record:
            existing_records.append(previous_record)
            allocation_lookup[applicant["applicant_number"]] = generated_values_from_previous(previous_record)
            continue

        next_new_number += 1
        applicant["applicant_number"] = next_new_number
        applicant["group"] = "Group 1" if next_new_number % 2 == 1 else "Group 2"
        new_applicants.append(applicant)

    existing_usage = usage_from_existing(existing_records)
    group_applicants = {
        "Group 1": [applicant for applicant in new_applicants if applicant["group"] == "Group 1"],
        "Group 2": [applicant for applicant in new_applicants if applicant["group"] == "Group 2"],
    }
    group_results = {
        group: allocate_group(items, mentor_counts, initial_usage=existing_usage[group])
        for group, items in group_applicants.items()
    }

    applicant_lookup = {applicant["applicant_number"]: applicant for applicant in new_applicants}
    for result in group_results.values():
        warnings.extend(result["warnings"])
        for item in result["allocations"]:
            applicant = applicant_lookup[item["applicant_number"]]
            expanded = result["final_capacity_per_mentor"] > 5
            if item["fallback_slots"] and expanded:
                status = "Allocated with fallback and capacity expansion"
            elif item["fallback_slots"]:
                status = "Allocated with fallback"
            elif expanded:
                status = "Allocated after capacity expansion"
            else:
                status = "Allocated"
            allocation_lookup[item["applicant_number"]] = {
                "Grouping": applicant["group"],
                "Event Sequence": "Panel Discussion -> Mentoring" if applicant["group"] == "Group 1" else "Mentoring -> Panel Discussion",
                "Subdomain Allocation 1": item["allocation_1"],
                "Subdomain Allocation 2": item["allocation_2"],
                "Allocation Status": status,
                "Attending": f"Panel Discussion; {item['allocation_1']} Mentoring; {item['allocation_2']} Mentoring",
            }

    output_df = df.copy()
    output_df["Applicant Number"] = [
        allocation_lookup.get(applicant["applicant_number"], {}).get("Applicant Number", applicant["applicant_number"])
        for applicant in applicants
    ]
    for column in GENERATED_COLUMNS[1:]:
        output_df[column] = [allocation_lookup.get(applicant["applicant_number"], {}).get(column, "Unable to allocate") for applicant in applicants]

    analytics = build_analytics(applicants, total_mentors, mentor_counts, group_results, max_applicants, previous_df, existing_records, new_applicants)
    analytics["capacity_warnings"] = build_existing_capacity_warnings(existing_usage, mentor_counts)
    return output_df, analytics, warnings


def build_analytics(applicants, total_mentors, mentor_counts, group_results, max_applicants, previous_df=None, existing_records=None, new_applicants=None):
    existing_records = [] if existing_records is None else existing_records
    new_applicants = applicants if new_applicants is None else new_applicants
    groups = {}
    for group, result in group_results.items():
        existing_count = sum(1 for record in existing_records if str(record.get("Grouping", "")).strip() == group)
        new_count = sum(1 for applicant in new_applicants if applicant["group"] == group)
        count = existing_count + new_count
        utilisation = {}
        for domain in DOMAIN_ORDER:
            capacity = result["capacity"][domain]
            utilisation[domain] = round((result["usage"][domain] / capacity) * 100, 1) if capacity else 0
        groups[group] = {
            "count": count,
            "required_allocations": count * 2,
            "final_capacity_per_mentor": result["final_capacity_per_mentor"],
            "usage": result["usage"],
            "capacity": result["capacity"],
            "remaining": result["remaining"],
            "utilisation": utilisation,
        }
    return {
        "total_applicants": len(applicants),
        "max_applicants": max_applicants,
        "within_maximum": len(applicants) <= max_applicants,
        "total_mentors": total_mentors,
        "mentor_counts": mentor_counts,
        "initial_capacity_per_mentor": 5,
        "groups": groups,
        "total_fallback_allocations": sum(result["fallback_count"] for result in group_results.values()),
        "total_unsuccessful_applicants": sum(result["unallocated_count"] for result in group_results.values()),
        "previous_allocation_rows": len(previous_df) if previous_df is not None else 0,
        "matched_existing_applicants": len(existing_records),
        "new_applicants_allocated": len(new_applicants),
    }


def make_summary_rows(analytics):
    return [
        ("Total applicants processed", analytics["total_applicants"]),
        ("Configured maximum applicants", analytics["max_applicants"]),
        ("Group 1 participant count", analytics["groups"]["Group 1"]["count"]),
        ("Group 2 participant count", analytics["groups"]["Group 2"]["count"]),
        ("Total mentors", analytics["total_mentors"]),
        ("Cybersecurity mentors", analytics["mentor_counts"]["Cybersecurity"]),
        ("Quantum Computing mentors", analytics["mentor_counts"]["Quantum Computing"]),
        ("Cloud Computing mentors", analytics["mentor_counts"]["Cloud Computing"]),
        ("Artificial Intelligence mentors", analytics["mentor_counts"]["Artificial Intelligence"]),
        ("Initial participants per mentor", analytics["initial_capacity_per_mentor"]),
        ("Group 1 final participants per mentor", analytics["groups"]["Group 1"]["final_capacity_per_mentor"]),
        ("Group 2 final participants per mentor", analytics["groups"]["Group 2"]["final_capacity_per_mentor"]),
        ("Previous allocation rows read", analytics["previous_allocation_rows"]),
        ("Existing applicants matched", analytics["matched_existing_applicants"]),
        ("New applicants allocated", analytics["new_applicants_allocated"]),
        ("Capacity warnings", len(analytics.get("capacity_warnings", []))),
        ("Unallocated applicants", analytics["total_unsuccessful_applicants"]),
    ]


def build_workbook(output_df, analytics, warnings):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        output_df.to_excel(writer, sheet_name="Allocated Applicants", index=False)
        pd.DataFrame(make_summary_rows(analytics), columns=["Summary Field", "Value"]).to_excel(
            writer, sheet_name="Allocation Summary", index=False
        )
        format_workbook(writer.book)
    output.seek(0)
    return output


def format_workbook(workbook):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    fallback_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F4CCCC")
    warning_fill = PatternFill("solid", fgColor="FCE4D6")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        headers = {cell.column: str(cell.value or "") for cell in sheet[1]}
        for row in sheet.iter_rows(min_row=2):
            values = [str(cell.value or "") for cell in row]
            fill = None
            if "Unable to allocate" in values:
                fill = error_fill
            elif any("fallback" in value.lower() for value in values):
                fill = fallback_fill
            elif sheet.title == "Warnings":
                fill = warning_fill
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill:
                    cell.fill = fill
                if "mobile" in headers[cell.column].lower() or "membership" in headers[cell.column].lower():
                    cell.number_format = "@"
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 45)


@app.route("/")
def index():
    return render_template("index.html", domains=DOMAIN_ORDER)


@app.post("/api/inspect")
def inspect_workbook():
    try:
        sheets, selected_sheet, df = load_sheet(request.files.get("file"), request.form.get("worksheet") or None)
        return jsonify(
            {
                "worksheets": sheets,
                "selectedWorksheet": selected_sheet,
                "columns": list(df.columns),
                "suggestedMappings": detect_option_columns(df.columns),
                "applicantCount": len(df),
                "previewRows": df.head(5).astype(str).to_dict(orient="records"),
            }
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/inspect-previous")
def inspect_previous_workbook():
    try:
        return jsonify(inspect_previous_allocation(request.files.get("previousFile"), request.form.get("previousWorksheet") or None))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/process")
def process_workbook():
    try:
        mappings = json.loads(request.form.get("mappings", "{}"))
        mentor_counts = json.loads(request.form.get("mentorCounts", "{}"))
        mentor_counts = {domain: int(mentor_counts.get(domain, 0)) for domain in DOMAIN_ORDER}
        max_applicants = int(request.form.get("maxApplicants", 400))
        total_mentors = int(request.form.get("totalMentors", 30))
        _, _, df = load_sheet(request.files.get("file"), request.form.get("worksheet") or None)
        previous_df = load_previous_allocations(request.files.get("previousFile"), request.form.get("previousWorksheet") or None)
        output_df, analytics, warnings = process_allocation(df, mappings, max_applicants, total_mentors, mentor_counts, previous_df)
        workbook = build_workbook(output_df, analytics, warnings)
        filename = f"mentoring_allocation_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
        response = send_file(
            workbook,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
        response.headers["X-Allocation-Analytics"] = base64.b64encode(json.dumps(analytics).encode("utf-8")).decode("ascii")
        return response
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        return jsonify({"error": "Processing failed. Please check the workbook and settings, then try again."}), 500


if __name__ == "__main__":
    app.run(debug=True)
