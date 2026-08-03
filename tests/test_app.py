from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from api.index import (
    DOMAIN_ORDER,
    GENERATED_COLUMNS,
    allocate_group,
    build_applicants,
    detect_option_columns,
    normalize_preference,
    process_allocation,
    validate_config,
    app,
)


MAPPINGS = {"first": "1st Option", "second": "2nd Option", "third": "3rd Option"}
MENTORS = {
    "Cybersecurity": 2,
    "Quantum Computing": 2,
    "Cloud Computing": 2,
    "Artificial Intelligence": 2,
}


def fake_df(count=4, prefs=None):
    prefs = prefs or [
        ("Cybersecurity", "Quantum Computing", "Cloud Computing"),
        ("Quantum Computing", "Cloud Computing", "Artificial Intelligence"),
        ("Cloud Computing", "Artificial Intelligence", "Cybersecurity"),
        ("Artificial Intelligence", "Cybersecurity", "Quantum Computing"),
    ]
    rows = []
    for index in range(count):
        first, second, third = prefs[index % len(prefs)]
        rows.append(
            {
                "Surname": f"Tester{index + 1}",
                "Given Name": f"Applicant{index + 1}",
                "Mobile Number": f"0987{index + 1:04d}",
                "Personal Email": f"applicant{index + 1}@example.com",
                "Membership Number": f"00{index + 1:04d}",
                "1st Option": first,
                "2nd Option": second,
                "3rd Option": third,
            }
        )
    return pd.DataFrame(rows)


def workbook_bytes(sheet_frames):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet, df in sheet_frames.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
    output.seek(0)
    return output


def test_normalises_reasonable_preference_variations():
    assert normalize_preference("AI") == "Artificial Intelligence"
    assert normalize_preference("Artificial intelligence") == "Artificial Intelligence"
    assert normalize_preference("Cloud computing") == "Cloud Computing"
    assert normalize_preference("Cyber security") == "Cybersecurity"
    assert normalize_preference("Quantum computing") == "Quantum Computing"


def test_detects_option_columns():
    columns = ["First Choice", "Second Option", "3rdOption"]
    assert detect_option_columns(columns) == {
        "first": "First Choice",
        "second": "Second Option",
        "third": "3rdOption",
    }


def test_empty_workbook_is_rejected():
    with pytest.raises(ValueError, match="does not contain any applicants"):
        process_allocation(pd.DataFrame(columns=MAPPINGS.values()), MAPPINGS, 400, 8, MENTORS)


def test_missing_option_columns_are_rejected():
    with pytest.raises(ValueError, match="map all three"):
        process_allocation(fake_df().drop(columns=["3rd Option"]), MAPPINGS, 400, 8, MENTORS)


def test_one_applicant_receives_two_distinct_allocations():
    output, analytics, warnings = process_allocation(fake_df(1), MAPPINGS, 400, 8, MENTORS)
    assert output.loc[0, "Grouping"] == "Group 1"
    assert output.loc[0, "Subdomain Allocation 1"] != output.loc[0, "Subdomain Allocation 2"]
    assert analytics["total_unsuccessful_applicants"] == 0
    assert warnings == []


def test_odd_and_even_grouping_preserves_order():
    output, _, _ = process_allocation(fake_df(5), MAPPINGS, 400, 8, MENTORS)
    assert output["Applicant Number"].tolist() == [1, 2, 3, 4, 5]
    assert output["Grouping"].tolist() == ["Group 1", "Group 2", "Group 1", "Group 2", "Group 1"]


def test_normal_top_two_allocation():
    output, _, _ = process_allocation(fake_df(1), MAPPINGS, 400, 8, MENTORS)
    assert output.loc[0, "Subdomain Allocation 1"] == "Cybersecurity"
    assert output.loc[0, "Subdomain Allocation 2"] == "Quantum Computing"


def test_first_choice_at_capacity_uses_second_choice():
    applicants, _ = build_applicants(fake_df(6, [("Cybersecurity", "Quantum Computing", "Cloud Computing")]), MAPPINGS)
    result = allocate_group(applicants, {"Cybersecurity": 1, "Quantum Computing": 2, "Cloud Computing": 2, "Artificial Intelligence": 2})
    sixth = result["allocations"][5]
    assert sixth["allocation_1"] == "Quantum Computing"


def test_second_choice_at_capacity_uses_third_choice():
    prefs = [("Cybersecurity", "Quantum Computing", "Cloud Computing")]
    applicants, _ = build_applicants(fake_df(6, prefs), MAPPINGS)
    result = allocate_group(applicants, {"Cybersecurity": 2, "Quantum Computing": 1, "Cloud Computing": 2, "Artificial Intelligence": 2})
    sixth = result["allocations"][5]
    assert sixth["allocation_2"] == "Cloud Computing"


def test_all_ranked_preferences_unavailable_uses_available_subdomain():
    prefs = [("Cybersecurity", "Quantum Computing", "Cloud Computing")]
    applicants, _ = build_applicants(fake_df(6, prefs), MAPPINGS)
    result = allocate_group(applicants, {"Cybersecurity": 1, "Quantum Computing": 1, "Cloud Computing": 1, "Artificial Intelligence": 2})
    assert result["allocations"][5]["allocation_2"] == "Artificial Intelligence"
    assert result["fallback_count"] >= 1


def test_duplicate_preferences_are_ignored():
    applicants, warnings = build_applicants(fake_df(1, [("AI", "AI", "Cloud Computing")]), MAPPINGS)
    assert applicants[0]["preferences"] == ["Artificial Intelligence", "Cloud Computing"]
    assert warnings[0]["Warning Type"] == "Duplicate preference"


def test_invalid_blank_and_no_valid_preferences_fall_back():
    df = fake_df(1, [("", "Unknown", "")])
    output, analytics, warnings = process_allocation(df, MAPPINGS, 400, 8, MENTORS)
    assert output.loc[0, "Subdomain Allocation 1"] in DOMAIN_ORDER
    assert output.loc[0, "Subdomain Allocation 2"] in DOMAIN_ORDER
    assert analytics["total_fallback_allocations"] == 2
    assert {warning["Warning Type"] for warning in warnings} >= {"Blank preference", "Invalid preference", "Fallback used"}


def test_exactly_400_applicants_allowed():
    output, analytics, _ = process_allocation(fake_df(400), MAPPINGS, 400, 8, MENTORS)
    assert len(output) == 400
    assert analytics["within_maximum"] is True


def test_more_than_400_applicants_rejected():
    with pytest.raises(ValueError, match="above the configured maximum"):
        process_allocation(fake_df(401), MAPPINGS, 400, 8, MENTORS)


def test_custom_maximum_applicant_limit():
    with pytest.raises(ValueError, match="above the configured maximum"):
        process_allocation(fake_df(11), MAPPINGS, 10, 8, MENTORS)


def test_uneven_mentor_distribution_and_zero_mentor_subdomain():
    mentors = {"Cybersecurity": 0, "Quantum Computing": 1, "Cloud Computing": 3, "Artificial Intelligence": 4}
    output, analytics, _ = process_allocation(fake_df(8), MAPPINGS, 400, 8, mentors)
    assert "Cybersecurity" not in output["Subdomain Allocation 1"].tolist() + output["Subdomain Allocation 2"].tolist()
    assert analytics["mentor_counts"]["Cloud Computing"] == 3


def test_only_one_subdomain_with_mentors_rejected():
    with pytest.raises(ValueError, match="At least two"):
        validate_config(400, 3, {"Cybersecurity": 3, "Quantum Computing": 0, "Cloud Computing": 0, "Artificial Intelligence": 0})


@pytest.mark.parametrize("count,expected", [(6, 6), (7, 7)])
def test_capacity_expands_as_needed(count, expected):
    applicants, _ = build_applicants(fake_df(count), MAPPINGS)
    result = allocate_group(applicants, {"Cybersecurity": 1, "Quantum Computing": 1, "Cloud Computing": 0, "Artificial Intelligence": 0})
    assert result["final_capacity_per_mentor"] == expected


def test_groups_can_have_different_final_capacity_limits():
    mentors = {"Cybersecurity": 1, "Quantum Computing": 1, "Cloud Computing": 0, "Artificial Intelligence": 0}
    _, analytics, _ = process_allocation(fake_df(13), MAPPINGS, 400, 2, mentors)
    assert analytics["groups"]["Group 1"]["final_capacity_per_mentor"] == 7
    assert analytics["groups"]["Group 2"]["final_capacity_per_mentor"] == 6


def test_capacity_is_reused_between_groups():
    mentors = {"Cybersecurity": 1, "Quantum Computing": 1, "Cloud Computing": 0, "Artificial Intelligence": 0}
    _, analytics, _ = process_allocation(fake_df(10), MAPPINGS, 400, 2, mentors)
    assert analytics["groups"]["Group 1"]["capacity"]["Cybersecurity"] == 5
    assert analytics["groups"]["Group 2"]["capacity"]["Cybersecurity"] == 5


def test_no_capacity_limit_exceeded_and_allocations_are_distinct():
    output, analytics, _ = process_allocation(fake_df(50), MAPPINGS, 400, 8, MENTORS)
    assert (output["Subdomain Allocation 1"] != output["Subdomain Allocation 2"]).all()
    for group in ["Group 1", "Group 2"]:
        for domain in DOMAIN_ORDER:
            assert analytics["groups"][group]["usage"][domain] <= analytics["groups"][group]["capacity"][domain]


def test_deterministic_repeated_output():
    first, first_analytics, _ = process_allocation(fake_df(30), MAPPINGS, 400, 8, MENTORS)
    second, second_analytics, _ = process_allocation(fake_df(30), MAPPINGS, 400, 8, MENTORS)
    assert first[GENERATED_COLUMNS].equals(second[GENERATED_COLUMNS])
    assert first_analytics == second_analytics


def test_original_columns_are_preserved_and_generated_columns_added():
    source = fake_df(2)
    output, _, _ = process_allocation(source, MAPPINGS, 400, 8, MENTORS)
    assert list(output.columns[: len(source.columns)]) == list(source.columns)
    for column in GENERATED_COLUMNS:
        assert column in output.columns


def test_invalid_mentor_total_rejected():
    with pytest.raises(ValueError, match="must equal total mentors"):
        validate_config(400, 9, MENTORS)


def test_completely_blank_rows_are_not_counted_through_inspect():
    client = app.test_client()
    df = pd.concat([fake_df(2), pd.DataFrame([{}])], ignore_index=True).fillna("")
    data = {"file": (workbook_bytes({"Registrations": df}), "registrations.xlsx")}
    response = client.post("/api/inspect", data=data, content_type="multipart/form-data")
    assert response.status_code == 200
    assert response.get_json()["applicantCount"] == 2


def test_multiple_worksheet_selection():
    client = app.test_client()
    data = {
        "file": (
            workbook_bytes({"Wrong Sheet": fake_df(1), "Registrations": fake_df(3)}),
            "registrations.xlsx",
        ),
        "worksheet": "Registrations",
    }
    response = client.post("/api/inspect", data=data, content_type="multipart/form-data")
    assert response.status_code == 200
    assert response.get_json()["selectedWorksheet"] == "Registrations"
    assert response.get_json()["applicantCount"] == 3


def test_successful_excel_download_response():
    client = app.test_client()
    data = {
        "file": (workbook_bytes({"Registrations": fake_df(4)}), "registrations.xlsx"),
        "worksheet": "Registrations",
        "mappings": json_dumps(MAPPINGS),
        "maxApplicants": "400",
        "totalMentors": "8",
        "mentorCounts": json_dumps(MENTORS),
    }
    response = client.post("/api/process", data=data, content_type="multipart/form-data")
    assert response.status_code == 200
    assert response.headers["Content-Type"].startswith("application/vnd.openxmlformats-officedocument")
    assert response.headers["X-Allocation-Analytics"]
    workbook = load_workbook(BytesIO(response.data))
    assert workbook.sheetnames == ["Allocated Applicants", "Allocation Summary"]


def test_generated_workbook_rejected_as_current_upload_in_inspect():
    client = app.test_client()
    allocated, _, _ = process_allocation(fake_df(2), MAPPINGS, 400, 8, MENTORS)
    response = client.post(
        "/api/inspect",
        data={"file": (workbook_bytes({"Allocated Applicants": allocated}), "allocated.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 400
    assert "generated allocation workbook" in response.get_json()["error"]


def test_generated_workbook_rejected_as_current_upload_in_process():
    client = app.test_client()
    allocated, _, _ = process_allocation(fake_df(2), MAPPINGS, 400, 8, MENTORS)
    data = {
        "file": (workbook_bytes({"Allocated Applicants": allocated}), "allocated.xlsx"),
        "worksheet": "Allocated Applicants",
        "mappings": json_dumps(MAPPINGS),
        "maxApplicants": "400",
        "totalMentors": "8",
        "mentorCounts": json_dumps(MENTORS),
    }
    response = client.post("/api/process", data=data, content_type="multipart/form-data")
    assert response.status_code == 400
    assert "generated allocation workbook" in response.get_json()["error"]


def test_previous_allocation_inspect_returns_worksheet_and_count():
    client = app.test_client()
    allocated, analytics, warnings = process_allocation(fake_df(4), MAPPINGS, 400, 8, MENTORS)
    workbook = workbook_bytes({"Allocated Applicants": allocated})
    response = client.post(
        "/api/inspect-previous",
        data={"previousFile": (workbook, "previous.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    data = response.get_json()
    assert data["selectedWorksheet"] == "Allocated Applicants"
    assert data["applicantCount"] == 4
    assert data["hasGeneratedColumns"] is True
    assert data["snapshot"]["groups"]["Group 1"]["participants"] == 2
    assert set(data["snapshot"]["groups"]["Group 1"]["usage"]) == set(DOMAIN_ORDER)


def test_previous_allocations_are_preserved_for_matched_applicants():
    first_batch, _, _ = process_allocation(fake_df(3), MAPPINGS, 400, 8, MENTORS)
    new_upload = fake_df(4)
    new_upload.loc[3, "Mobile Number"] = "099999999"
    new_upload.loc[3, "Personal Email"] = "new@example.com"

    output, analytics, _ = process_allocation(new_upload, MAPPINGS, 400, 8, MENTORS, first_batch)

    assert output.loc[0, "Subdomain Allocation 1"] == first_batch.loc[0, "Subdomain Allocation 1"]
    assert output.loc[0, "Subdomain Allocation 2"] == first_batch.loc[0, "Subdomain Allocation 2"]
    assert analytics["matched_existing_applicants"] == 3
    assert analytics["new_applicants_allocated"] == 1


def test_matched_existing_generated_fields_are_ported_exactly():
    first_batch, _, _ = process_allocation(fake_df(1), MAPPINGS, 400, 8, MENTORS)
    first_batch.loc[0, "Applicant Number"] = 88
    first_batch.loc[0, "Grouping"] = "Group 2"
    first_batch.loc[0, "Event Sequence"] = "Mentoring -> Panel Discussion"
    first_batch.loc[0, "Subdomain Allocation 1"] = "Artificial Intelligence"
    first_batch.loc[0, "Subdomain Allocation 2"] = "Cloud Computing"
    first_batch.loc[0, "Allocation Status"] = "Manual preserved status"
    first_batch.loc[0, "Attending"] = "Panel Discussion; Artificial Intelligence Mentoring; Cloud Computing Mentoring"

    output, analytics, _ = process_allocation(fake_df(1), MAPPINGS, 400, 8, MENTORS, first_batch)

    for column in GENERATED_COLUMNS:
        assert str(output.loc[0, column]) == str(first_batch.loc[0, column])
    assert analytics["matched_existing_applicants"] == 1
    assert analytics["new_applicants_allocated"] == 0


def test_previous_allocations_do_not_match_when_mobile_or_email_changes():
    first_batch, _, _ = process_allocation(fake_df(1), MAPPINGS, 400, 8, MENTORS)
    new_upload = fake_df(1)
    new_upload.loc[0, "Mobile Number"] = "099999999"
    new_upload.loc[0, "Personal Email"] = "changed@example.com"

    output, analytics, _ = process_allocation(new_upload, MAPPINGS, 400, 8, MENTORS, first_batch)

    assert output.loc[0, "Allocation Status"] == "Allocated"
    assert analytics["matched_existing_applicants"] == 0
    assert analytics["new_applicants_allocated"] == 1


def test_previous_rows_are_source_of_truth_after_manual_removal():
    first_batch, _, _ = process_allocation(fake_df(5), MAPPINGS, 400, 8, MENTORS)
    manually_trimmed = first_batch.head(3).copy()
    new_upload = fake_df(4)
    new_upload.loc[3, "Mobile Number"] = "099999999"
    new_upload.loc[3, "Personal Email"] = "new@example.com"

    output, analytics, _ = process_allocation(new_upload, MAPPINGS, 400, 8, MENTORS, manually_trimmed)

    assert analytics["previous_allocation_rows"] == 3
    assert analytics["matched_existing_applicants"] == 3
    assert output.loc[3, "Applicant Number"] == 4


def test_previous_only_rows_are_appended_to_final_output():
    previous_allocated, _, _ = process_allocation(fake_df(3), MAPPINGS, 400, 8, MENTORS)
    current_upload = fake_df(2)

    output, analytics, _ = process_allocation(current_upload, MAPPINGS, 400, 8, MENTORS, previous_allocated)

    assert len(output) == 3
    assert analytics["matched_existing_applicants"] == 2
    assert analytics["previous_rows_appended"] == 1
    assert output.iloc[2]["Personal Email"] == previous_allocated.iloc[2]["Personal Email"]
    assert output.iloc[2]["Subdomain Allocation 1"] == previous_allocated.iloc[2]["Subdomain Allocation 1"]


def test_decreased_capacity_warns_without_changing_existing_allocations():
    previous = fake_df(12, [("Cybersecurity", "Quantum Computing", "Cloud Computing")])
    previous_allocated, _, _ = process_allocation(previous, MAPPINGS, 400, 8, MENTORS)
    new_upload = previous.copy()
    reduced_mentors = {
        "Cybersecurity": 1,
        "Quantum Computing": 1,
        "Cloud Computing": 0,
        "Artificial Intelligence": 0,
    }

    output, analytics, _ = process_allocation(new_upload, MAPPINGS, 400, 2, reduced_mentors, previous_allocated)

    assert output["Subdomain Allocation 1"].tolist() == previous_allocated["Subdomain Allocation 1"].tolist()
    assert analytics["capacity_warnings"]


def test_invalid_file_extension_rejected_by_inspect():
    client = app.test_client()
    data = {"file": (BytesIO(b"not excel"), "registrations.csv")}
    response = client.post("/api/inspect", data=data, content_type="multipart/form-data")
    assert response.status_code == 400


def test_event_sequence_values_are_generated():
    output, _, _ = process_allocation(fake_df(2), MAPPINGS, 400, 8, MENTORS)
    assert output.loc[0, "Event Sequence"] == "Panel Discussion -> Mentoring"
    assert output.loc[1, "Event Sequence"] == "Mentoring -> Panel Discussion"


def test_attending_column_contains_panel_and_two_mentorings():
    output, _, _ = process_allocation(fake_df(1), MAPPINGS, 400, 8, MENTORS)
    attending = output.loc[0, "Attending"]
    assert "Panel Discussion" in attending
    assert attending.count("Mentoring") == 2


def test_summary_analytics_include_required_capacity_fields():
    _, analytics, _ = process_allocation(fake_df(4), MAPPINGS, 400, 8, MENTORS)
    assert analytics["groups"]["Group 1"]["required_allocations"] == 4
    assert set(analytics["groups"]["Group 1"]["capacity"]) == set(DOMAIN_ORDER)
    assert set(analytics["groups"]["Group 2"]["remaining"]) == set(DOMAIN_ORDER)


def test_same_column_cannot_be_used_for_multiple_mappings():
    bad_mappings = {"first": "1st Option", "second": "1st Option", "third": "3rd Option"}
    with pytest.raises(ValueError, match="different Excel column"):
        process_allocation(fake_df(2), bad_mappings, 400, 8, MENTORS)


def json_dumps(value):
    import json

    return json.dumps(value)
